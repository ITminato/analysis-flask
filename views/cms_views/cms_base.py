# -*- coding: utf-8 -*-
import csv, os, datetime
from threading import Thread
from openpyxl import Workbook
from flask import views, request, session, redirect, url_for, abort, render_template, current_app
from . import bp
from common_utils import xtjson
from common_utils.upload import UploadCls
from models.cms_user import CmsUserModel
from common_utils.utils_funcs import PagingCLS
from ..view_func import add_admin_log, cms_risk_control
from common_utils.utils_funcs import generate_filename
from constants import CMS_USER_SESSION_KEY, OnClick, SITE_CONFIG_CACHE, CMS_MENUS_GROUPS, EventType, ExportStatu, EXPORT_FOLDER, IMAGES_TYPES, ASSETS_FOLDER, PermissionType
from models.customer_table import ExportDataModel


@bp.before_request
def site_cms_before_request():
    res = cms_risk_control()
    if res:
        return res


class CmsViewBase(views.MethodView):

    title = ''
    template = ''
    endpoint = ''
    xtjson = xtjson
    permission_map = {}
    MCLS = None
    add_url_rules = [[]]

    def __init_data(self):
        self.user_uuid = session.get(CMS_USER_SESSION_KEY)
        self.current_admin_user = CmsUserModel.query_one({'uuid': self.user_uuid})
        self.is_superdamin = False
        if self.current_admin_user:
            self.is_superdamin = self.current_admin_user.is_superadmin
        self.EventType = EventType
        self.data_from = {}
        self.filter_dict = {}
        self.context = {}
        self.project_name = current_app.config.get('PROJECT_NAME')
        self.project_static_folder = os.path.join(current_app._static_folder, self.project_name)
        self.context['current_admin_user'] = self.current_admin_user
        self.context['site_data'] = SITE_CONFIG_CACHE
        self.context['load_menus'] = self.load_menus
        self.context['cms_menu_groups'] = CMS_MENUS_GROUPS
        self.context['FIELDS'] = self.MCLS.fields()
        self.context['is_superdamin'] = self.is_superdamin

    @classmethod
    def html_StringField(cls, db_field, field_cls, data_dict={}):
        html = '<div class="input-group mb-3">'
        html += '<div class="input-group-prepend">'
        html += '<span class="input-group-text">'
        if not field_cls.nullable:
            html += f'<span class="text-danger">*</span>{field_cls.field_name}：</span></div>'
        else:
            html += f'{field_cls.field_name}：</span></div>'
        if data_dict.get(db_field):
            html += u'<input type="text" class="form-control" id="%s" placeholder="%s" value="%s">' % (db_field, field_cls.placeholder, data_dict.get(db_field))
        else:
            html += u'<input type="text" class="form-control " id="%s" placeholder="%s">' % (db_field, field_cls.placeholder)
        html += u'</div>'
        return html


    def load_menus(self, datas=[]):
        if not datas:
            return []
        p_menus = []
        for menu in datas:
            if self.check_permission('%s&%s' % (menu.get('endpoint'), PermissionType.ACCESS)):
                p_menus.append(menu)
        return p_menus

    def check_login(self):
        if not self.current_admin_user:
            self.addcmslog_func(EventType.ACCESS, '未登录！跳转登录页面')
            return redirect(url_for('admin.cms_login'))
        if not self.current_admin_user.statu:
            session.pop(CMS_USER_SESSION_KEY)
            return abort(404)


    def check_permission(self, code):
        if self.is_superdamin:
            return True
        if not self.current_admin_user or not code:
            return False
        return self.current_admin_user.has_permission(code)

    def check_superdamin(self, pers):
        if 'superadmin' in pers:
            return True
        return

    def addcmslog_func(self, event_type, msg):
        if self.current_admin_user:
            add_admin_log(event_type, msg, admin_uuid=self.current_admin_user.uuid)
        else:
            add_admin_log(event_type, msg)

    def search_func(self, FIELDS):
        """get数据搜索处理"""
        s_filter_dict, s_context_res = {}, {}
        if hasattr(self.MCLS, 'field_search'):
            field_search = getattr(self.MCLS, 'field_search')()
            for db_field in field_search:
                col_value = request.args.get(db_field)
                if col_value is None:
                    continue
                s_context_res[db_field] = col_value
                if col_value and col_value.strip():
                    col_value = col_value.strip()
                    field_cls = FIELDS.get(db_field)
                    if not field_cls:
                        return  False, '%s: 无处理属性!' % db_field
                    if field_cls.field_type == 'UUIDField':
                        s_filter_dict[db_field] = col_value
                    elif field_cls.field_type == 'IDField':
                        s_filter_dict[db_field] = int(col_value)
                    else:
                        statu, res = field_cls.search_validate(col_value)
                        if not statu:
                            return False, res
                        s_filter_dict[db_field] = res
        return True, [s_filter_dict, s_context_res]

    def export_xlsx_func(self, uuid, export_folder, filename, datas, headers, filead_datas=[]):
        if not os.path.exists(export_folder):
            os.makedirs(export_folder)

        export_data = ExportDataModel.find_one({'uuid': uuid}) or {}
        if not export_data:
            return

        try:
            crr_count = 0
            wb = Workbook()
            wa = wb.active
            row = 1
            for h in range(len(headers)):
                wa.cell(row=row, column=h + 1, value=headers[h])

            for data in datas:
                row += 1
                for index, filed in enumerate(filead_datas):
                    _v = data.get(filed.get('db_filed'))
                    if filed.get('type') == 'int':
                        wa.cell(row=row, column=index+1, value=int(_v or 0))
                    elif filed.get('type') == 'float':
                        wa.cell(row=row, column=index+1, value=float(_v or 0))
                    elif filed.get('type') == 'datetime':
                        if _v and isinstance(_v, datetime.datetime):
                            wa.cell(row=row, column=index+1, value=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                    else:
                        wa.cell(row=row, column=index + 1, value=_v or '')
                crr_count += 1
                if crr_count % 100 == 0:
                    export_data['out_count'] = crr_count
                    ExportDataModel.save(export_data)

            file_path = os.path.join(export_folder, filename)
            wb.save(file_path)
            export_data['out_count'] = crr_count
            export_data['statu'] = ExportStatu.successed
            ExportDataModel.save(export_data)
            return True
        except Exception as e:
            export_data['note'] = str(e)
            export_data['statu'] = ExportStatu.failed
            ExportDataModel.save(export_data)
            return

    def editorUploadFile(self, foldername='article/images'):
        """编辑器上传文件处理"""
        if not self.check_permission('%s&%s'%(self.endpoint, PermissionType.UPLOAD_DATA)):
            data = { "uploaded": 0, "error": { "message": '无文件操作权限!'}}
            return xtjson.json_result(**data)
        upload = UploadCls()
        upload.static_folder = self.project_static_folder
        upload.uploaddir = ASSETS_FOLDER
        upload.foldername = foldername
        upload.limit_types = IMAGES_TYPES
        statu, msg = upload.upload_file_func()
        if statu:
            filename = msg.rsplit('/', 1)[-1]
            result_dict = {"uploaded": 1, "fileName": filename, "url": msg}
        else:
            result_dict = {"uploaded": 0, "error": { "message": msg}}
        return xtjson.json_result(**result_dict)

    def post_upload_picture(self, picture_name='', foldername='article/images', limit_types=IMAGES_TYPES, limit_size=100):
        """
        图片上传处理
        :param picture_name: 图片名称
        :param uploaddir: 上传到的文件夹名称(1级)
        :param foldername:  上传到的文件父级文件夹名称(2级)
        :param limit_size: 文件大小限制(KB)
        :param limit_types:
        :return:
        """
        if not self.check_permission('%s&%s'%(self.endpoint, PermissionType.UPLOAD_DATA)):
            return self.no_permission()
        upload = UploadCls()
        upload.static_folder = self.project_static_folder
        upload.uploaddir = ASSETS_FOLDER
        upload.foldername = foldername
        upload.limit_types = limit_types
        upload.filename = picture_name
        statu, msg = upload.upload_file_func()
        return statu, msg

    def post_file_upload(self, filename='', foldername='', limit_types=IMAGES_TYPES):
        """
        文件上传处理
        :param filename: 文件名
        :param uploaddir: 上传到的文件夹名称(1级)
        :param foldername:  上传到的文件父级文件夹名称(2级)
        :param limmit_types:  文件类型限制
        """
        if not self.check_permission('%s_%s'%(self.endpoint, PermissionType.UPLOAD_DATA)):
            return self.no_permission()
        upload = UploadCls()
        upload.static_folder = self.project_static_folder
        upload.uploaddir = ASSETS_FOLDER
        upload.foldername = foldername
        upload.limit_types = limit_types
        upload.filename = filename
        return upload.upload_file_func()

    def has_permission_func(self):
        """权限检测"""
        operation_type = ('_add_', '_del_', '_edit_', '_out_', '_upload_')
        if not self.check_permission('%s_%s'%(self.endpoint, PermissionType.ACCESS)):
            return
        for oper in operation_type:
            if self.action.startswith(oper):
                if oper == '_add_':
                    crr_p = '%s&%s'%(self.endpoint, PermissionType.ADD)
                    if not self.check_permission(crr_p):
                        return
                if oper == '_del_':
                    crr_p = '%s&%s'%(self.endpoint, PermissionType.DELETE)
                    if not self.check_permission(crr_p):
                        return
                if oper == '_edit_':
                    crr_p = '%s&%s'%(self.endpoint, PermissionType.EDIT)
                    if not self.check_permission(crr_p):
                        return
                if oper == '_out_':
                    crr_p = '%s&%s'%(self.endpoint, PermissionType.EXPORT_DATA)
                    if not self.check_permission(crr_p):
                        return
                if oper == '_upload_':
                    crr_p = '%s&%s'%(self.endpoint, PermissionType.UPLOAD_DATA)
                    if not self.check_permission(crr_p):
                        return
        return True

    def no_permission(self):
        return self.xtjson.json_params_error('无操作权限!')

    def get_other_way(self):
        return

    def get_context(self):
        """获取context内容"""
        return {}

    def get_filter_dict(self):
        """获取搜索参数"""
        return {}

    def post_data_other_way(self):
        return

    def post_other_way(self):
        return

    def view_get(self, *args, **kwargs):
        return self.xtjson.json_result()

    def view_post(self, *args, **kwargs):
        return self.xtjson.json_result()

    def get(self, *args, **kwargs):
        self.__init_data()
        res = self.check_login()
        if res:
            return res
        add_admin_log(EventType.ACCESS)
        return self.view_get(*args, **kwargs)

    def post(self, *args, **kwargs):
        self.__init_data()
        res = self.check_login()
        if res:
            return res
        return self.view_post(*args, **kwargs)


class CmsFormViewBase(CmsViewBase):
    title = ''
    template = ''
    show_menu = True
    MCLS = CmsUserModel



